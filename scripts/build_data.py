#!/usr/bin/env python3
"""
build_data.py — Converte Site/Bufalo_Catalogo_Estruturado.xlsx
para src/data/categorias.json e src/data/produtos.json
"""

import json
import os
import re
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(BASE_DIR, "Site", "Bufalo_Catalogo_Estruturado.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "src", "data")

os.makedirs(DATA_DIR, exist_ok=True)

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ── Categorias ──────────────────────────────────────────────────────────────
ws_cat = wb["Categorias"]
categorias = []
slug_by_nome = {}

for row in ws_cat.iter_rows(min_row=2, values_only=True):
    id_cat, nome, slug, descricao, ordem = row
    if id_cat is None:
        continue
    categorias.append({
        "id": str(id_cat).strip(),
        "nome": str(nome).strip(),
        "slug": str(slug).strip(),
        "descricao": str(descricao).strip() if descricao else "",
        "ordem": int(ordem) if ordem is not None else 0,
    })
    slug_by_nome[str(nome).strip()] = str(slug).strip()

# ── Produtos ─────────────────────────────────────────────────────────────────
ws_prod = wb["Produtos"]
produtos = []

for row in ws_prod.iter_rows(min_row=2, values_only=True):
    id_prod, categoria, nome, composicao, medida, embalagem, cores, destaque_raw, obs = row
    if id_prod is None:
        continue

    categoria_str = str(categoria).strip() if categoria else ""
    categoria_slug = slug_by_nome.get(categoria_str, "")
    destaque = str(destaque_raw).strip().lower() == "sim" if destaque_raw else False

    produto = {
        "id": str(id_prod).strip(),
        "categoria": categoria_str,
        "categoriaSlug": categoria_slug,
        "nome": str(nome).strip() if nome else "",
        "composicao": str(composicao).strip() if composicao else "",
        "medida": str(medida).strip() if medida else "",
        "embalagem": str(embalagem).strip() if embalagem else "",
        "cores": str(cores).strip() if cores else "",
        "destaque": destaque,
        "obs": str(obs).strip() if obs else "",
        "imagem": f"/images/produtos/{str(id_prod).strip()}.webp",
    }
    produtos.append(produto)

# ── Salvar JSONs ─────────────────────────────────────────────────────────────
cat_path = os.path.join(DATA_DIR, "categorias.json")
prod_path = os.path.join(DATA_DIR, "produtos.json")

with open(cat_path, "w", encoding="utf-8") as f:
    json.dump(categorias, f, ensure_ascii=False, indent=2)

with open(prod_path, "w", encoding="utf-8") as f:
    json.dump(produtos, f, ensure_ascii=False, indent=2)

print(f"categorias.json: {len(categorias)} itens")
print(f"produtos.json:   {len(produtos)} itens")

assert len(categorias) == 8, f"Esperado 8 categorias, obtido {len(categorias)}"
assert len(produtos) == 31,  f"Esperado 31 produtos, obtido {len(produtos)}"
print("OK 8 31")
